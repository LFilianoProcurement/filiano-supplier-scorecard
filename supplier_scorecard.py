import streamlit as st
import json
import csv
import io
import os
import datetime
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

st.set_page_config(
    page_title="Supplier Scorecard Tool",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

load_dotenv()

# ── CSS ────────────────────────────────────────────────────
st.markdown("""
<style>
body { background-color: #F3F4F6; color: #111827; }
.stApp { background-color: #F3F4F6; }
.stTextArea textarea { border: 2px solid #1F4E79 !important; border-radius: 6px !important; }
.stTextArea textarea:focus { border: 2px solid #2E75B6 !important; box-shadow: 0 0 0 1px #2E75B6 !important; }
[data-testid="stSidebar"] { background-color: #FFFFFF; border-right: 2px solid #1F4E79; }
[data-testid="stSidebar"] label { color: #111827 !important; font-weight: 600; }
[data-testid="stSidebar"] p { color: #111827 !important; }
[data-testid="stSidebar"] div { color: #111827 !important; }
.stTextInput input { border: 2px solid #1F4E79 !important; color: #111827 !important; background: #FFFFFF !important; }
.stSelectbox [data-baseweb="select"] { border: 2px solid #1F4E79 !important; }
.stButton button { background-color: #1F4E79 !important; color: #FFFFFF !important; font-weight: 700 !important; border: none !important; }
.stButton button:hover { background-color: #2E75B6 !important; }
.stButton button p { color: #FFFFFF !important; }
.stSlider label { color: #111827 !important; font-weight: 600 !important; }
h1,h2,h3 { color: #1F4E79; }
p, div, span, label { color: #111827; }
[data-testid="stMarkdownContainer"] p { color: #111827 !important; }
.stTabs [data-baseweb="tab"] { color: #374151 !important; font-weight: 600; }
.stTabs [aria-selected="true"] { color: #1F4E79 !important; border-bottom: 3px solid #1F4E79; }

.score-card {
    background: #FFFFFF;
    border: 1px solid #E5E7EB;
    border-radius: 10px;
    padding: 20px;
    text-align: center;
    box-shadow: 0 2px 4px rgba(0,0,0,0.06);
    margin-bottom: 12px;
}
.score-number {
    font-size: 3rem;
    font-weight: 800;
    line-height: 1;
    margin-bottom: 4px;
}
.score-label {
    font-size: 0.75rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    color: #6B7280;
}
.category-card {
    background: #FFFFFF;
    border: 1px solid #E5E7EB;
    border-radius: 8px;
    padding: 16px;
    margin-bottom: 10px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05);
}
.category-title {
    font-size: 1rem;
    font-weight: 700;
    color: #1F4E79;
    margin-bottom: 8px;
}
.sub-element-row {
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding: 6px 0;
    border-bottom: 1px solid #F3F4F6;
    font-size: 0.85rem;
}
.badge-gold { background: #FEF3C7; color: #92400E; border: 1px solid #FCD34D; padding: 3px 10px; border-radius: 12px; font-size: 0.72rem; font-weight: 700; }
.badge-silver { background: #F3F4F6; color: #374151; border: 1px solid #D1D5DB; padding: 3px 10px; border-radius: 12px; font-size: 0.72rem; font-weight: 700; }
.badge-bronze { background: #FEF3C7; color: #78350F; border: 1px solid #F59E0B; padding: 3px 10px; border-radius: 12px; font-size: 0.72rem; font-weight: 700; }
.badge-needs-improvement { background: #FEE2E2; color: #991B1B; border: 1px solid #FCA5A5; padding: 3px 10px; border-radius: 12px; font-size: 0.72rem; font-weight: 700; }
.legend-bar {
    background: #FFFFFF;
    border: 1px solid #E5E7EB;
    border-radius: 8px;
    padding: 12px 20px;
    margin-bottom: 16px;
    display: flex;
    gap: 20px;
    align-items: center;
    flex-wrap: wrap;
}
.title-block {
    padding: 16px 0;
    border-bottom: 2px solid #1F4E79;
    margin-bottom: 20px;
}
.na-badge { background: #F3F4F6; color: #6B7280; border: 1px solid #D1D5DB; padding: 3px 8px; border-radius: 12px; font-size: 0.72rem; }
#MainMenu { visibility: hidden; }
footer { visibility: hidden; }
header { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


# ── SCORECARD STRUCTURE ────────────────────────────────────
SCORECARD_CATEGORIES = {
    "1 - Quality": {
        "weight": 25,
        "sub_elements": {
            "1.1 - Audit Observations & FCAs": {
                "desc": "FDA Warning Letters, Form 483s, FCAs, regulatory actions from external audits",
                "score_1": "At least 1 Warning Letter or 2+ regulatory actions from audits",
                "score_2": "No Warning Letter but at least 1 Form 483 or FCA",
                "score_3": "No Warning Letters/483s/FCAs but 1 regulatory action from audits",
                "score_4": "No Warning Letters/483s/FCAs, no regulatory actions, meeting quality targets",
                "score_5": "No Warning Letters/483s/FCAs, no regulatory actions, exceeding quality targets",
                "source": "FDA Form 483 Database, FDA Warning Letter Database"
            },
            "1.2 - Lot Acceptance Rate (LAR)": {
                "desc": "Percentage of lots accepted on first submission",
                "score_1": "LAR < 80%",
                "score_2": "LAR > 80%",
                "score_3": "LAR > 90%",
                "score_4": "LAR > 95%",
                "score_5": "LAR is 100%",
                "source": "Supplier Quality Dashboard - LAR% Tab"
            },
            "1.3 - Continuous Improvement": {
                "desc": "CI program infrastructure, Six Sigma/Lean deployment, resource dedication",
                "score_1": "CI program not developed, isolated effort only",
                "score_2": "CI program developed and being implemented, key personnel identified",
                "score_3": "CI deployed with defined roles, 0.5% dedicated resources, Lean BBs exist",
                "score_4": "CI fully deployed, certified GB/BB/Lean BB team, MBB exists",
                "score_5": "Mature CI, supervisors and operators own CI projects, rotational development",
                "source": "Supplier Management Self-Assessment Survey"
            },
            "1.4 - SCARs": {
                "desc": "Supplier Corrective Action Requests in last 12 months",
                "score_1": "4 or more SCARs in last 12 months",
                "score_2": "3 SCARs in last 12 months",
                "score_3": "2 SCARs in last 12 months",
                "score_4": "1 SCAR in last 12 months",
                "score_5": "Zero SCARs in last 12 months",
                "source": "Supplier Quality Dashboard - SCAR Volume Tab"
            },
            "1.5 - NCRs": {
                "desc": "Non-Conformance Reports in last 12 months",
                "score_1": "4 or more NCRs in last 12 months",
                "score_2": "3 NCRs in last 12 months",
                "score_3": "2 NCRs in last 12 months",
                "score_4": "1 NCR in last 12 months",
                "score_5": "Zero NCRs in last 12 months",
                "source": "Supplier Quality Dashboard - NCR Volume Tab"
            },
        }
    },
    "2 - Delivery": {
        "weight": 15,
        "sub_elements": {
            "2.1 - On Time Delivery (OTD) & Predicted vs. Actual": {
                "desc": "OTD percentage and forecast accuracy",
                "score_1": "OTD < 90% & PvsA < 80%",
                "score_2": "OTD 90-94% & PvsA 80-84%",
                "score_3": "OTD 94-96% & PvsA 85-89%",
                "score_4": "OTD 96-99% & PvsA 90-94%",
                "score_5": "OTD > 99% & PvsA > 95%",
                "source": "Supplier Quality Dashboard - OTD Tab"
            },
            "2.2 - Capacity Status": {
                "desc": "Available capacity headroom above current demand",
                "score_1": "0-5% headroom for growth lines; 0-3% for stable lines",
                "score_2": "6-14% headroom for growth lines; 4-7% for stable lines",
                "score_3": "15-20% headroom for growth lines; 8-9% for stable lines",
                "score_4": "21-29% headroom for growth lines; 10-14% for stable lines",
                "score_5": "30%+ headroom for growth lines; 15%+ for stable lines",
                "source": "Commodity Manager assessment"
            },
            "2.3 - Dock to Stock": {
                "desc": "Percentage of items on dock-to-stock certified supplier program",
                "score_1": "Dock-to-stock on < 10% of items",
                "score_2": "Dock-to-stock on 10-40% of items",
                "score_3": "Dock-to-stock on 40-60% of items",
                "score_4": "Dock-to-stock on 60-80% of items",
                "score_5": "Dock-to-stock on 80-100% of items",
                "source": "Certified Supplier List"
            },
        }
    },
    "3 - Cost": {
        "weight": 15,
        "sub_elements": {
            "3.1 - Cost Management": {
                "desc": "Cost reduction programs and 3-year cost trend",
                "score_1": "No cost reduction program, upward cost trend over 3 years",
                "score_2": "Cost reduction program exists but upward cost trend continues",
                "score_3": "Cost reduction program exists with flat normalized cost trend",
                "score_4": "Cost reduction program with >1% supplier-driven reduction annually",
                "score_5": "Cost sharing program exists between supplier and buyer",
                "source": "Contracts / ERP system assessment"
            },
            "3.2 - Cost of Quality": {
                "desc": "Field Corrective Actions (FCAs) frequency",
                "score_1": "1 FCA in last 12 months",
                "score_2": "1 FCA in last 2 years",
                "score_3": "1 FCA in last 3 years",
                "score_4": "1 FCA in last 4 years",
                "score_5": "No FCAs in last 5 years",
                "source": "Supplier Quality Dashboard - SCAR Volume Tab"
            },
        }
    },
    "4 - Execution & Responsiveness": {
        "weight": 20,
        "sub_elements": {
            "4.1 - Notification of Change (NOC) Timeliness": {
                "desc": "Proactive notification of changes affecting product form, fit, or function",
                "score_1": "Did not notify of a change impacting product in last 2 years",
                "score_2": "Notified but without ample time to implement the change",
                "score_3": "Notified with ample time but no safety stock or implementation plan provided",
                "score_4": "Notified with ample time, provided safety stock and implementation schedule",
                "score_5": "Notified with ample time, safety stock, schedule, and proactive collaboration",
                "source": "NOC Database manual assessment"
            },
            "4.2 - SCAR Aging": {
                "desc": "Timeliness of SCAR implementation phase completion",
                "score_1": "1+ SCAR completed > 90 days past due date or > 3 extensions",
                "score_2": "1+ SCAR completed 15-90 days past due date or > 2 extensions",
                "score_3": "All SCARs on time but 1+ extension made",
                "score_4": "All SCARs on time with zero extensions",
                "score_5": "All SCARs on time, zero extensions, supplier owned closure without SQE oversight",
                "source": "SCAR phase completion dates in QMS"
            },
            "4.3 - Schedule / Project Adherence": {
                "desc": "Commitment reliability and follow-through without escalation",
                "score_1": "Commitments go overdue, supplier does not drive actions to completion",
                "score_2": "Commitments overdue, completion only with multiple follow-ups",
                "score_3": "Occasional overdue, requires moderate follow-up",
                "score_4": "Mostly on time, requires only minimal follow-up",
                "score_5": "Consistently on time, zero follow-up required, self-managing",
                "source": "Historical project performance assessment"
            },
        }
    },
    "5 - Inventory": {
        "weight": 10,
        "sub_elements": {
            "5.1 - Vendor Managed Inventory (VMI)": {
                "desc": "VMI or consignment inventory coverage on A-items",
                "score_1": "VMI on < 40% of A items",
                "score_2": "VMI on > 40% of A items",
                "score_3": "Consignment inventory on < 40% of A items",
                "score_4": "Consignment inventory on > 40% and < 70% of A items",
                "score_5": "Consignment inventory on > 70% of A items",
                "source": "Contracts / ERP / JDA assessment"
            },
            "5.2 - Inventory Plans & Material Shortages": {
                "desc": "Frequency and management of material shortages (Contract Manufacturers)",
                "score_1": "More than 15 shortage issues in 3 months",
                "score_2": "10 or more issues in 3 months, shortages known weekly",
                "score_3": "4-10 shortages in 3 months, weekly tracking and expediting",
                "score_4": "1-3 shortages in 3 months, shortages are exceptions",
                "score_5": "No shortages in a 3-month period",
                "source": "ERP / Contracts assessment"
            },
        }
    },
    "6 - Business Continuity": {
        "weight": 10,
        "sub_elements": {
            "6.1 - Business Continuity Plan (BCP)": {
                "desc": "BCP documentation and implementation level",
                "score_1": "BCP not documented",
                "score_2": "BCP 25% implemented",
                "score_3": "BCP 50% implemented",
                "score_4": "BCP 75% implemented",
                "score_5": "BCP 100% implemented",
                "source": "Supplier Management Self-Assessment Survey"
            },
            "6.2 - Sustainability": {
                "desc": "Environmental, social, and financial sustainability plan",
                "score_1": "Sustainability plan not documented",
                "score_2": "Sustainability plan 25% implemented",
                "score_3": "Sustainability plan 50% implemented",
                "score_4": "Sustainability plan 75% implemented",
                "score_5": "Sustainability plan 100% implemented",
                "source": "Supplier Management Self-Assessment Survey"
            },
            "6.3 - Financial Health / Stability": {
                "desc": "FRISK® (public) or PAYCE® (private) credit risk score",
                "score_1": "FRISK®/PAYCE® score of 1-3 (high bankruptcy risk)",
                "score_2": "FRISK®/PAYCE® score of 4-5",
                "score_3": "FRISK®/PAYCE® score of 6-7",
                "score_4": "FRISK®/PAYCE® score of 8-9",
                "score_5": "FRISK®/PAYCE® score of 10 (lowest risk)",
                "source": "CreditRiskMonitor.com"
            },
        }
    },
    "7 - Innovation": {
        "weight": 5,
        "sub_elements": {
            "7.1 - Strategic Alignment": {
                "desc": "Joint innovation roadmap development and investment",
                "score_1": "No plans to develop an Innovation Roadmap",
                "score_2": "Plan exists to develop an Innovation Roadmap",
                "score_3": "Moderate development of an Innovation Roadmap underway",
                "score_4": "Material investment allocated, roadmap clearly defined and jointly prioritized",
                "score_5": "Significant investment, mature roadmap with proven concepts aligned to business needs",
                "source": "SRT team manual assessment"
            },
            "7.2 - Innovation Capacity": {
                "desc": "Dedicated resources and infrastructure to support innovation",
                "score_1": "No infrastructure exists to support innovation",
                "score_2": "Capabilities exist to support innovation",
                "score_3": "Capabilities and infrastructure in place",
                "score_4": "Dedicated resources and stable infrastructure",
                "score_5": "Dedicated, highly-qualified resources with mature innovation infrastructure",
                "source": "SRT team manual assessment"
            },
            "7.3 - Cultural Alignment": {
                "desc": "Transparency, information sharing, and partnership depth",
                "score_1": "All information confidential, no sharing",
                "score_2": "Projects identified for information sharing",
                "score_3": "Limited transparency, partnership limited to specific projects",
                "score_4": "Moderate transparency and information sharing across multiple projects",
                "score_5": "Supplier is extension of buyer, respectful transparency, shared IP exists",
                "source": "SRT team manual assessment"
            },
        }
    },
}

THRESHOLDS = {"Gold": 0.90, "Silver": 0.75, "Bronze": 0.60}


def get_tier(pct):
    if pct >= THRESHOLDS["Gold"]:
        return "Gold"
    elif pct >= THRESHOLDS["Silver"]:
        return "Silver"
    elif pct >= THRESHOLDS["Bronze"]:
        return "Bronze"
    else:
        return "Needs Improvement"


def get_tier_color(tier):
    colors = {
        "Gold": "#D97706",
        "Silver": "#6B7280",
        "Bronze": "#B45309",
        "Needs Improvement": "#DC2626"
    }
    return colors.get(tier, "#6B7280")


def get_score_color(pct):
    if pct >= 0.90:
        return "#16A34A"
    elif pct >= 0.75:
        return "#2E75B6"
    elif pct >= 0.60:
        return "#D97706"
    else:
        return "#DC2626"


def calculate_scores(scores):
    """Calculate scores based only on scored elements.
    Unscored and N/A items are excluded — the score auto-adjusts to
    reflect only the elements actually evaluated."""
    category_results = {}
    total_weighted = 0
    total_weight = 0

    for cat_name, cat_data in SCORECARD_CATEGORIES.items():
        actual_score = 0
        scored_count = 0
        na_count = 0
        total_items = len(cat_data["sub_elements"])

        for sub_name in cat_data["sub_elements"].keys():
            val = scores.get(f"{cat_name}||{sub_name}", None)
            if val == "N/A":
                na_count += 1
            elif val is not None:
                actual_score += val
                scored_count += 1

        # Max possible = only the items actually scored (5 pts each)
        max_possible = scored_count * 5

        if max_possible > 0:
            pct = actual_score / max_possible
        else:
            pct = 0

        not_scored = total_items - scored_count - na_count

        category_results[cat_name] = {
            "actual": actual_score,
            "max": max_possible,
            "pct": pct,
            "tier": get_tier(pct) if scored_count > 0 else "Not Scored",
            "scored_items": scored_count,
            "na_items": na_count,
            "not_scored": not_scored,
            "total_items": total_items
        }

        # Only include category in overall if at least one item was scored
        if scored_count > 0:
            weight = cat_data["weight"]
            total_weighted += pct * weight
            total_weight += weight

    overall_pct = total_weighted / total_weight if total_weight > 0 else 0
    overall_tier = get_tier(overall_pct) if total_weight > 0 else "Not Scored"

    return category_results, overall_pct, overall_tier


RESPONSES_FILE = "survey_responses.json"

def save_to_unified_dashboard(supplier_name, reviewer, period, category_results, overall_pct):
    """Save scorecard results to shared survey_responses.json for unified dashboard view"""
    import datetime, json, os
    # Build category scores in same format as customer survey
    cat_scores = {}
    for cat_name, result in category_results.items():
        if result["scored_items"] > 0:
            cat_scores[cat_name] = {"avg": round(result["pct"] * 5, 2), "source": "internal"}

    record = {
        "id": datetime.datetime.now().strftime("%Y%m%d%H%M%S%f"),
        "submitted_at": datetime.datetime.now().isoformat(),
        "supplier": supplier_name,
        "customer_name": reviewer,
        "customer_company": "Internal — Procurement",
        "overall_avg": round(overall_pct * 5, 2),
        "scores": cat_scores,
        "comments": {},
        "source": "internal",
        "period": period,
        "weight": 3
    }
    existing = []
    try:
        if os.path.exists(RESPONSES_FILE):
            with open(RESPONSES_FILE, "r") as f:
                existing = json.load(f)
    except:
        pass
    # Replace any existing internal record for same supplier+period
    existing = [r for r in existing if not (
        r.get("source") == "internal" and
        r.get("supplier", "").lower() == supplier_name.lower() and
        r.get("period", "") == period
    )]
    existing.append(record)
    with open(RESPONSES_FILE, "w") as f:
        json.dump(existing, f, indent=2)
    return True


def scores_to_csv(supplier_name, reviewer, period, scores, category_results, overall_pct, overall_tier):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["SUPPLIER SCORECARD REPORT", "", "", ""])
    writer.writerow(["Supplier", supplier_name, "Reviewer", reviewer])
    writer.writerow(["Period", period, "Date", datetime.date.today().isoformat()])
    writer.writerow(["Overall Score", f"{overall_pct:.1%}", "Overall Tier", overall_tier])
    writer.writerow([])
    writer.writerow(["Category", "Sub-Element", "Score", "Max", "N/A"])
    for cat_name, cat_data in SCORECARD_CATEGORIES.items():
        for sub_name in cat_data["sub_elements"].keys():
            key = f"{cat_name}||{sub_name}"
            val = scores.get(key, None)
            is_na = val == "N/A"
            writer.writerow([cat_name, sub_name,
                              "" if is_na or val is None else val,
                              5 if not is_na else "N/A",
                              "Yes" if is_na else "No"])
        res = category_results[cat_name]
        writer.writerow([cat_name + " TOTAL", "", res["actual"], res["max"],
                         f"{res['pct']:.1%} - {res['tier']}"])
        writer.writerow([])
    return output.getvalue()





# ══════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════
def make_bar_chart(category_results, supplier_name, period):
    """Generate a bar chart and return as PNG bytes"""
    scored = {k: v for k, v in category_results.items() if v["tier"] != "Not Scored" and v["max"] > 0}
    if not scored:
        return None
    labels = [k.split(" - ")[1] if " - " in k else k for k in scored.keys()]
    values = [v["pct"] * 100 for v in scored.values()]
    colors = []
    for v in scored.values():
        if v["pct"] >= 0.90: colors.append("#16A34A")
        elif v["pct"] >= 0.75: colors.append("#2E75B6")
        elif v["pct"] >= 0.60: colors.append("#D97706")
        else: colors.append("#DC2626")
    fig, ax = plt.subplots(figsize=(10, max(4, len(labels) * 0.7)))
    fig.patch.set_facecolor("#FFFFFF")
    ax.set_facecolor("#F9FAFB")
    bars = ax.barh(labels, values, color=colors, height=0.55, edgecolor="white", linewidth=0.5)
    ax.axvline(x=90, color="#16A34A", linestyle="--", linewidth=1.2, alpha=0.7, label="Gold 90%")
    ax.axvline(x=75, color="#2E75B6", linestyle="--", linewidth=1.2, alpha=0.7, label="Silver 75%")
    ax.axvline(x=60, color="#D97706", linestyle="--", linewidth=1.2, alpha=0.7, label="Bronze 60%")
    for bar, val in zip(bars, values):
        ax.text(min(val + 1, 97), bar.get_y() + bar.get_height()/2,
                f"{val:.0f}%", va="center", ha="left", fontsize=9, fontweight="bold", color="#1F2937")
    ax.set_xlim(0, 105)
    ax.set_xlabel("Score (%)", fontsize=10, color="#374151")
    ax.set_title(f"{supplier_name}  |  {period}", fontsize=12, fontweight="bold", color="#1F4E79", pad=12)
    ax.tick_params(colors="#374151", labelsize=9)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#E5E7EB")
    ax.spines["bottom"].set_color("#E5E7EB")
    ax.legend(loc="lower right", fontsize=8, framealpha=0.9)
    ax.invert_yaxis()
    plt.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="#FFFFFF")
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def make_excel_report(supplier_name, reviewer, period, scores, category_results, overall_pct, overall_tier):
    """Generate Excel report with summary, detail scores, and embedded bar chart"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Scorecard Summary"
    navy, blue = "1F4E79", "2E75B6"
    thin = Side(style="thin", color="D1D5DB")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "SUPPLIER SCORECARD REPORT"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    # Info rows
    for r, (l1, v1, l2, v2) in enumerate([
        ("Supplier", supplier_name, "Reviewer", reviewer),
        ("Period", period, "Date", datetime.date.today().isoformat()),
        ("Overall Score", f"{overall_pct:.1%}", "Overall Tier", overall_tier),
    ], 2):
        ws.cell(r, 1, l1).font = Font(name="Arial", bold=True, size=10, color=navy)
        ws.cell(r, 2, v1).font = Font(name="Arial", size=10)
        ws.cell(r, 4, l2).font = Font(name="Arial", bold=True, size=10, color=navy)
        ws.cell(r, 5, v2).font = Font(name="Arial", size=10)
        ws.row_dimensions[r].height = 18
    # Category table header
    sr = 6
    for c, h in enumerate(["Category", "Scored Pts", "Max Pts", "Score %", "Tier", "Elements Scored"], 1):
        cell = ws.cell(sr, c, h)
        cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.alignment = Alignment(horizontal="center")
        cell.border = bdr
    tier_colors = {"Gold": "FEF3C7", "Silver": "F3F4F6", "Bronze": "FEF9C3",
                   "Needs Improvement": "FEE2E2", "Not Scored": "F9FAFB"}
    for r2, (cat, res) in enumerate(category_results.items(), sr + 1):
        bg = tier_colors.get(res["tier"], "FFFFFF")
        scored_note = f"{res['scored_items']} of {res['total_items']}"
        if res.get("na_items", 0) > 0:
            scored_note += f" ({res['na_items']} N/A)"
        for c2, val in enumerate([cat, res["actual"], res["max"],
                                    f"{res['pct']:.1%}" if res["max"] > 0 else "N/A",
                                    res["tier"], scored_note], 1):
            cell = ws.cell(r2, c2, val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = bdr
            if c2 > 1:
                cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[r2].height = 16
    ov = sr + len(category_results) + 1
    ws.merge_cells(f"A{ov}:C{ov}")
    ws.cell(ov, 1, "OVERALL SCORE").font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws.cell(ov, 1).fill = PatternFill("solid", fgColor=navy)
    ws.cell(ov, 4, f"{overall_pct:.1%}").font = Font(name="Arial", bold=True, size=11, color=navy)
    ws.cell(ov, 4).alignment = Alignment(horizontal="center")
    ws.cell(ov, 5, overall_tier).font = Font(name="Arial", bold=True, size=11, color=navy)
    ws.cell(ov, 5).alignment = Alignment(horizontal="center")
    ws.row_dimensions[ov].height = 20
    # Chart data and chart
    scored_cats = [(k, v) for k, v in category_results.items() if v["tier"] != "Not Scored" and v["max"] > 0]
    if scored_cats:
        cd = ov + 3
        ws.cell(cd, 1, "Category").font = Font(name="Arial", bold=True, size=9)
        ws.cell(cd, 2, "Score %").font = Font(name="Arial", bold=True, size=9)
        for i, (cat, res) in enumerate(scored_cats, cd + 1):
            short = cat.split(" - ")[1] if " - " in cat else cat
            ws.cell(i, 1, short)
            ws.cell(i, 2, round(res["pct"] * 100, 1))
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = f"{supplier_name} - Category Scores ({period})"
        chart.x_axis.title = "Category"
        chart.y_axis.title = "Score (%)"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100
        chart.style = 10
        chart.width = 24
        chart.height = 14
        chart.add_data(Reference(ws, min_col=2, min_row=cd, max_row=cd + len(scored_cats)), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=cd + 1, max_row=cd + len(scored_cats)))
        # Add threshold reference lines as series
        from openpyxl.chart import Series
        from openpyxl.chart.data_source import NumDataSource, NumRef
        ws.add_chart(chart, f"A{cd + len(scored_cats) + 2}")
    # Detail sheet
    ws2 = wb.create_sheet("Detail Scores")
    for c, h in enumerate(["Category", "Sub-Element", "Score", "N/A"], 1):
        ws2.cell(1, c, h).font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        ws2.cell(1, c).fill = PatternFill("solid", fgColor=navy)
    dr = 2
    for cat_name, cat_data in SCORECARD_CATEGORIES.items():
        for sub_name in cat_data["sub_elements"].keys():
            val = scores.get(f"{cat_name}||{sub_name}", None)
            is_na = val == "N/A"
            ws2.cell(dr, 1, cat_name).font = Font(name="Arial", size=9)
            ws2.cell(dr, 2, sub_name).font = Font(name="Arial", size=9)
            ws2.cell(dr, 3, "" if is_na or val is None else val).font = Font(name="Arial", size=9)
            ws2.cell(dr, 4, "Yes" if is_na else ("No" if val is not None else "")).font = Font(name="Arial", size=9)
            if is_na:
                ws2.cell(dr, 3).fill = PatternFill("solid", fgColor="F3F4F6")
            dr += 1
    # Column widths
    ws.column_dimensions["A"].width = 36
    for col, w in [("B",12),("C",12),("D",12),("E",20),("F",18)]:
        ws.column_dimensions[col].width = w
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 42
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 8
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()



def main():

    # Sidebar
    with st.sidebar:
        st.markdown("## 📊 Supplier Scorecard")
        st.markdown("*Procurement Intelligence Suite*")
        st.markdown("---")
        supplier_name = st.text_input("Supplier Name", placeholder="e.g. Sterigenics")
        reviewer = st.text_input("Your Name / Team", placeholder="e.g. Louis Filiano")
        period = st.selectbox("Review Period",
                               ["Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025",
                                "Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026",
                                "Annual 2024", "Annual 2025", "Annual 2026"])
        st.markdown("---")
        st.markdown("""
**Scoring Scale:**
- **5** = Exceeds expectations
- **4** = Meets expectations
- **3** = Approaching expectations
- **2** = Below expectations
- **1** = Significantly below

**Tiers:**
- 🥇 Gold ≥ 90%
- 🥈 Silver ≥ 75%
- 🥉 Bronze ≥ 60%
- ⚠️ Needs Improvement < 60%
""")
        st.markdown("---")
        st.markdown("*Louis Filiano — Procurement Intelligence Suite*")

    # Header
    st.markdown("""
<div class="title-block">
    <div style="font-size:1.8rem; font-weight:700; color:#1F4E79;">📊 Supplier Scorecard Tool</div>
    <div style="font-size:0.9rem; color:#6B7280; margin-top:4px;">Performance evaluation framework for regulated manufacturing suppliers</div>
</div>
""", unsafe_allow_html=True)

    # Legend
    st.markdown("""
<div class="legend-bar">
    <div style="font-size:0.72rem; font-weight:700; color:#6B7280; text-transform:uppercase; letter-spacing:0.08em;">PERFORMANCE TIERS:</div>
    <div style="display:flex; align-items:center; gap:6px;">
        <span class="badge-gold">🥇 GOLD</span>
        <span style="font-size:0.8rem; color:#374151;">Score &ge; 90%</span>
    </div>
    <div style="display:flex; align-items:center; gap:6px;">
        <span class="badge-silver">🥈 SILVER</span>
        <span style="font-size:0.8rem; color:#374151;">Score &ge; 75%</span>
    </div>
    <div style="display:flex; align-items:center; gap:6px;">
        <span class="badge-bronze">🥉 BRONZE</span>
        <span style="font-size:0.8rem; color:#374151;">Score &ge; 60%</span>
    </div>
    <div style="display:flex; align-items:center; gap:6px;">
        <span class="badge-needs-improvement">⚠️ NEEDS IMPROVEMENT</span>
        <span style="font-size:0.8rem; color:#374151;">Score &lt; 60%</span>
    </div>
    <div style="border-left:1px solid #E5E7EB; padding-left:16px; font-size:0.8rem; color:#374151;">
        Score each sub-element 1-5 or mark N/A if not applicable
    </div>
</div>
""", unsafe_allow_html=True)

    if not supplier_name:
        st.info("👈 Enter a supplier name in the sidebar to begin.")
        return

    # ── SCORING TABS ───────────────────────────────────────
    tab_names = ["📝 Enter Scores"] + [f"{k.split(' - ')[0]}" for k in SCORECARD_CATEGORIES.keys()] + ["📊 Results", "📋 Action Items"]
    tabs = st.tabs(["📝 Score Entry", "1-Quality", "2-Delivery", "3-Cost", "4-Execution", "5-Inventory", "6-BCP", "7-Innovation", "📊 Results", "📋 Actions"])

    # Initialize session state for scores
    if "scores" not in st.session_state:
        st.session_state.scores = {}
    if "action_items" not in st.session_state:
        st.session_state.action_items = []

    # Auto-reset if supplier or period changes
    current_context = f"{supplier_name}||{period}"
    if st.session_state.get("_last_context", "") != current_context:
        if st.session_state.get("_last_context", ""):
            # Context changed — clear scores and sliders
            keys_to_delete = [k for k in st.session_state.keys()
                               if k.startswith("slider_") or k.startswith("na_")]
            for k in keys_to_delete:
                del st.session_state[k]
            st.session_state.scores = {}
            st.session_state.action_items = []
        st.session_state["_last_context"] = current_context

    # ── TAB 0: SCORE ENTRY OVERVIEW ────────────────────────
    with tabs[0]:
        st.markdown(f"### Scoring: **{supplier_name}** | Period: {period}")
        st.markdown("Click each category tab above to enter scores. Use this tab for a quick overview of completion status.")

        col1, col2, col3 = st.columns(3)
        total_elements = sum(len(v["sub_elements"]) for v in SCORECARD_CATEGORIES.values())
        scored = sum(1 for k, v in st.session_state.scores.items() if v is not None)

        with col1:
            st.markdown(f"""
<div class="score-card">
    <div class="score-number" style="color:#1F4E79;">{scored}</div>
    <div class="score-label">Elements Scored</div>
</div>""", unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
<div class="score-card">
    <div class="score-number" style="color:#6B7280;">{total_elements}</div>
    <div class="score-label">Total Elements</div>
</div>""", unsafe_allow_html=True)
        with col3:
            pct_complete = scored / total_elements if total_elements > 0 else 0
            st.markdown(f"""
<div class="score-card">
    <div class="score-number" style="color:#16A34A;">{pct_complete:.0%}</div>
    <div class="score-label">Complete</div>
</div>""", unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("### Category Completion Status")
        for cat_name, cat_data in SCORECARD_CATEGORIES.items():
            scored_cat = sum(1 for sub in cat_data["sub_elements"]
                            if st.session_state.scores.get(f"{cat_name}||{sub}") is not None
                            and st.session_state.scores.get(f"{cat_name}||{sub}") != "N/A")
            na_cat = sum(1 for sub in cat_data["sub_elements"]
                        if st.session_state.scores.get(f"{cat_name}||{sub}") == "N/A")
            total_cat = len(cat_data["sub_elements"])
            color = "#16A34A" if scored_cat + na_cat == total_cat else "#D97706" if scored_cat > 0 else "#6B7280"
            na_note = f", {na_cat} N/A" if na_cat > 0 else ""
            remaining = total_cat - scored_cat - na_cat
            remaining_note = f", {remaining} remaining" if remaining > 0 else ""
            st.markdown(f"""
<div style="display:flex; justify-content:space-between; align-items:center; padding:8px 12px; background:#FFFFFF; border-radius:6px; margin-bottom:6px; border:1px solid #E5E7EB;">
    <span style="font-weight:600; color:#1F4E79; font-size:0.9rem;">{cat_name}</span>
    <span style="color:{color}; font-weight:700; font-size:0.85rem;">{scored_cat} scored{na_note}{remaining_note}</span>
</div>""", unsafe_allow_html=True)

    # ── CATEGORY TABS 1-7 ──────────────────────────────────
    for tab_idx, (cat_name, cat_data) in enumerate(SCORECARD_CATEGORIES.items(), 1):
        with tabs[tab_idx]:
            st.markdown(f"### {cat_name}")
            st.markdown(f"*Weight: {cat_data['weight']}% of overall score*")
            st.markdown("---")

            for sub_name, sub_data in cat_data["sub_elements"].items():
                key = f"{cat_name}||{sub_name}"
                st.markdown(f"**{sub_name}**")
                st.markdown(f"<span style='color:#6B7280; font-size:0.83rem;'>{sub_data['desc']}</span>", unsafe_allow_html=True)

                col_score, col_na = st.columns([4, 1])
                with col_na:
                    is_na = st.checkbox("N/A", key=f"na_{key}")

                with col_score:
                    if is_na:
                        st.session_state.scores[key] = "N/A"
                        st.markdown('<span style="color:#6B7280; font-size:0.85rem;">Marked as Not Applicable</span>', unsafe_allow_html=True)
                    else:
                        current_val = st.session_state.scores.get(key)
                        if current_val == "N/A":
                            current_val = None

                        # Use "— Not Scored —" as first option so default is blank
                        score_val = st.select_slider(
                            f"Score for {sub_name}",
                            options=["— Not Scored —", 1, 2, 3, 4, 5],
                            value=current_val if current_val and current_val != "N/A" else "— Not Scored —",
                            key=f"slider_{key}",
                            label_visibility="collapsed"
                        )

                        if score_val == "— Not Scored —":
                            # Remove from scores so it counts as unscored
                            st.session_state.scores.pop(key, None)
                            st.markdown('<span style="color:#9CA3AF; font-size:0.82rem; font-style:italic;">Move slider to score this element</span>', unsafe_allow_html=True)
                        else:
                            st.session_state.scores[key] = score_val
                            # Show scoring criteria
                            criteria = sub_data.get(f"score_{score_val}", "")
                            if criteria:
                                st.markdown(f'<div style="background:#EFF6FF; border-left:3px solid #2E75B6; padding:6px 10px; border-radius:0 4px 4px 0; font-size:0.8rem; color:#1F2937; margin-top:4px;"><strong>Score {score_val}:</strong> {criteria}</div>', unsafe_allow_html=True)

                st.markdown(f'<span style="color:#9CA3AF; font-size:0.75rem;">Source: {sub_data["source"]}</span>', unsafe_allow_html=True)
                st.markdown("---")

    # ── TAB 8: RESULTS ────────────────────────────────────
    with tabs[8]:
        category_results, overall_pct, overall_tier = calculate_scores(st.session_state.scores)
        tier_color = get_tier_color(overall_tier)
        tier_emoji = {"Gold": "🥇", "Silver": "🥈", "Bronze": "🥉", "Needs Improvement": "⚠️"}

        st.markdown(f"### Results: **{supplier_name}** | {period}")

        # Overall score cards
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.markdown(f"""
<div class="score-card">
    <div class="score-number" style="color:{get_score_color(overall_pct)};">{overall_pct:.0%}</div>
    <div class="score-label">Overall Score</div>
    <div style="margin-top:8px;"><span class="badge-{overall_tier.lower().replace(' ', '-')}">{tier_emoji.get(overall_tier,'')} {overall_tier.upper()}</span></div>
</div>""", unsafe_allow_html=True)

        scored_count = sum(1 for v in st.session_state.scores.values() if v is not None and v != "N/A")
        with col2:
            st.markdown(f"""
<div class="score-card">
    <div class="score-number" style="color:#1F4E79;">{scored_count}</div>
    <div class="score-label">Elements Scored</div>
</div>""", unsafe_allow_html=True)

        gold_cats = sum(1 for r in category_results.values() if r["tier"] == "Gold")
        with col3:
            st.markdown(f"""
<div class="score-card">
    <div class="score-number" style="color:#D97706;">{gold_cats}</div>
    <div class="score-label">Gold Categories</div>
</div>""", unsafe_allow_html=True)

        needs_imp = sum(1 for r in category_results.values() if r["tier"] == "Needs Improvement")
        with col4:
            st.markdown(f"""
<div class="score-card">
    <div class="score-number" style="color:#DC2626;">{needs_imp}</div>
    <div class="score-label">Needs Improvement</div>
</div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # Category breakdown
        st.markdown("#### Category Breakdown")
        for cat_name, result in category_results.items():
            pct = result["pct"]
            tier = result["tier"]
            color = get_score_color(pct)
            bar_width = int(pct * 100)
            tier_badge = f'<span class="badge-{tier.lower().replace(" ", "-")}">{tier_emoji.get(tier,"")} {tier}</span>'

            # Build element count note
            scored_note = f"{result['scored_items']} of {result['total_items']} elements scored"
            if result['na_items'] > 0:
                scored_note += f", {result['na_items']} N/A"
            if result['not_scored'] > 0:
                scored_note += f", {result['not_scored']} not yet scored"
            if result['max'] > 0:
                pts_each = result['max'] / result['scored_items']
                scored_note += f" ({pts_each:.0f} pts each)"

            if result['tier'] == "Not Scored":
                st.markdown(f"""
<div class="category-card" style="opacity:0.6;">
    <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px;">
        <span class="category-title">{cat_name}</span>
        <span style="color:#6B7280; font-size:0.85rem;">Not yet scored</span>
    </div>
    <div style="font-size:0.78rem; color:#9CA3AF;">{scored_note}</div>
</div>""", unsafe_allow_html=True)
            else:
                st.markdown(f"""
<div class="category-card">
    <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px;">
        <span class="category-title">{cat_name}</span>
        <div style="display:flex; align-items:center; gap:12px;">
            {tier_badge}
            <span style="font-weight:700; color:{color}; font-size:1.1rem;">{pct:.0%}</span>
            <span style="color:#6B7280; font-size:0.82rem;">{result['actual']}/{result['max']} pts</span>
        </div>
    </div>
    <div style="background:#F3F4F6; border-radius:4px; height:10px;">
        <div style="width:{bar_width}%; background:{color}; border-radius:4px; height:10px;"></div>
    </div>
    <div style="font-size:0.78rem; color:#6B7280; margin-top:6px;">{scored_note}</div>
</div>""", unsafe_allow_html=True)

        # Threshold reference
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("""
<div style="background:#FFFFFF; border:1px solid #E5E7EB; border-radius:8px; padding:12px 16px; display:flex; gap:24px; align-items:center; flex-wrap:wrap;">
    <span style="font-size:0.75rem; font-weight:700; color:#6B7280; text-transform:uppercase;">THRESHOLDS:</span>
    <span style="font-size:0.83rem; color:#D97706; font-weight:700;">🥇 Gold ≥ 90%</span>
    <span style="font-size:0.83rem; color:#6B7280; font-weight:700;">🥈 Silver ≥ 75%</span>
    <span style="font-size:0.83rem; color:#B45309; font-weight:700;">🥉 Bronze ≥ 60%</span>
    <span style="font-size:0.83rem; color:#DC2626; font-weight:700;">⚠️ Needs Improvement &lt; 60%</span>
</div>
""", unsafe_allow_html=True)

        # Bar chart
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("#### Score Chart")
        chart_bytes = make_bar_chart(category_results, supplier_name, period)
        if chart_bytes:
            st.image(chart_bytes, use_container_width=True)
        else:
            st.info("Score at least one category to see the chart.")

        # Save to unified dashboard
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("#### Save to Unified Dashboard")
        if st.button("💾 Save Scorecard to Survey Dashboard", use_container_width=True, key="save_unified"):
            if supplier_name and category_results:
                success = save_to_unified_dashboard(supplier_name, reviewer, period,
                                                     category_results, overall_pct)
                if success:
                    st.success(f"✅ Scorecard for **{supplier_name}** saved to the unified dashboard! Open Survey Admin to see it alongside customer responses.")
                else:
                    st.error("Could not save — check that survey_responses.json is accessible.")
            else:
                st.warning("Enter a supplier name and score at least one category first.")

        # Download buttons
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("#### Download Report")
        dl_col1, dl_col2 = st.columns(2)
        with dl_col1:
            csv_data = scores_to_csv(supplier_name, reviewer, period,
                                      st.session_state.scores, category_results,
                                      overall_pct, overall_tier)
            st.download_button(
                label="📥 Download CSV Report",
                data=csv_data,
                file_name=f"scorecard_{supplier_name.replace(' ', '_')}_{period.replace(' ', '_')}.csv",
                mime="text/csv",
                use_container_width=True
            )
        with dl_col2:
            xlsx_data = make_excel_report(supplier_name, reviewer, period,
                                           st.session_state.scores, category_results,
                                           overall_pct, overall_tier)
            st.download_button(
                label="📊 Download Excel Report (with Chart)",
                data=xlsx_data,
                file_name=f"scorecard_{supplier_name.replace(' ', '_')}_{period.replace(' ', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        # Clear scorecard
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("#### Reset Scorecard")
        st.markdown('<p style="color:#6B7280; font-size:0.85rem;">Clear all scores to start fresh for a new QBR period. Downloads and saved dashboard data are not affected.</p>', unsafe_allow_html=True)
        confirm_reset = st.checkbox("Confirm — clear all scores for this scorecard", key="confirm_reset")
        if st.button("🗑️ Clear All Scores", use_container_width=True, key="clear_scores_btn"):
            if confirm_reset:
                # Clear all slider and checkbox widget states so tabs show blank
                keys_to_delete = [k for k in st.session_state.keys()
                                   if k.startswith("slider_") or k.startswith("na_")]
                for k in keys_to_delete:
                    del st.session_state[k]
                st.session_state.scores = {}
                st.session_state.action_items = []
                st.session_state.editing_item = None
                st.success("✅ Scorecard cleared — all tabs reset. Ready for next QBR period!")
                st.rerun()
            else:
                st.warning("Please check the confirmation box first.")

    # ── TAB 9: ACTION ITEMS ───────────────────────────────
    with tabs[9]:
        st.markdown(f"### Action Items: **{supplier_name}** | {period}")
        st.markdown("Track action items from the scorecard review meeting.")

        # Action item inputs outside form so values persist on submit
        col1, col2 = st.columns(2)
        with col1:
            action_cat = st.selectbox("Category", list(SCORECARD_CATEGORIES.keys()), key="ai_cat")
            st.markdown("**Action Item**")
            action_item = st.text_area("Action Item", height=80,
                                        placeholder="Describe the required action...",
                                        key="ai_item", label_visibility="collapsed")
            action_status = st.selectbox("Status", ["Open", "In Progress", "Completed", "On Hold"],
                                          key="ai_status")
        with col2:
            action_due = st.date_input("Due Date",
                                        value=datetime.date.today() + datetime.timedelta(days=30),
                                        key="ai_due")
            action_owner_buyer = st.text_input("Buyer Owner",
                                                placeholder="e.g. Louis Filiano", key="ai_buyer")
            action_owner_supplier = st.text_input("Supplier Owner",
                                                   placeholder="e.g. John Smith", key="ai_supplier")
        action_notes = st.text_area("Current Actions Taken / Comments", height=60, key="ai_notes")

        if st.button("➕ Add Action Item", use_container_width=True, key="ai_add_btn"):
            if st.session_state.ai_item.strip():
                st.session_state.action_items.append({
                    "category": st.session_state.ai_cat,
                    "action": st.session_state.ai_item,
                    "status": st.session_state.ai_status,
                    "due_date": st.session_state.ai_due.isoformat(),
                    "buyer_owner": st.session_state.ai_buyer,
                    "supplier_owner": st.session_state.ai_supplier,
                    "notes": st.session_state.ai_notes
                })
                st.success(f"Action item added! ({len(st.session_state.action_items)} total)")
                st.rerun()
            else:
                st.error("Action Item is required — please describe the action before adding.")

        if st.session_state.action_items:
            st.markdown("---")
            st.markdown(f"#### Current Action Items ({len(st.session_state.action_items)})")
            st.markdown('<span style="font-size:0.8rem; color:#6B7280;">Click <strong>Edit</strong> on any item to update its status, due date, or notes</span>', unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)

            # Track which item is being edited
            if "editing_item" not in st.session_state:
                st.session_state.editing_item = None

            for i, item in enumerate(st.session_state.action_items):
                status_color = {"Open": "#DC2626", "In Progress": "#D97706",
                                "Completed": "#16A34A", "On Hold": "#6B7280"}.get(item["status"], "#6B7280")

                is_editing = st.session_state.editing_item == i

                if is_editing:
                    # Inline edit mode
                    with st.container():
                        st.markdown(f'''<div style="background:#EFF6FF; border:2px solid #1F4E79; border-radius:8px; padding:14px; margin-bottom:8px;">''', unsafe_allow_html=True)
                        st.markdown(f"**Editing:** {item['action'][:80]}{'...' if len(item['action']) > 80 else ''}")
                        ec1, ec2, ec3 = st.columns(3)
                        with ec1:
                            new_status = st.selectbox("Status", ["Open", "In Progress", "Completed", "On Hold"],
                                                       index=["Open", "In Progress", "Completed", "On Hold"].index(item["status"]),
                                                       key=f"edit_status_{i}")
                        with ec2:
                            new_due = st.date_input("Due Date",
                                                     value=datetime.date.fromisoformat(item["due_date"]),
                                                     key=f"edit_due_{i}")
                        with ec3:
                            new_buyer = st.text_input("Buyer Owner", value=item["buyer_owner"], key=f"edit_buyer_{i}")
                        new_notes = st.text_area("Notes / Current Actions", value=item["notes"], height=70, key=f"edit_notes_{i}")
                        st.markdown("</div>", unsafe_allow_html=True)

                        col_save, col_cancel, col_delete = st.columns([2, 2, 1])
                        with col_save:
                            if st.button("💾 Save Changes", key=f"save_{i}", use_container_width=True):
                                st.session_state.action_items[i]["status"] = new_status
                                st.session_state.action_items[i]["due_date"] = new_due.isoformat()
                                st.session_state.action_items[i]["buyer_owner"] = new_buyer
                                st.session_state.action_items[i]["notes"] = new_notes
                                st.session_state.editing_item = None
                                st.rerun()
                        with col_cancel:
                            if st.button("✕ Cancel", key=f"cancel_{i}", use_container_width=True):
                                st.session_state.editing_item = None
                                st.rerun()
                        with col_delete:
                            if st.button("🗑️", key=f"delete_{i}", use_container_width=True, help="Delete this action item"):
                                st.session_state.action_items.pop(i)
                                st.session_state.editing_item = None
                                st.rerun()
                else:
                    # Display mode
                    col_item, col_btn = st.columns([6, 1])
                    with col_item:
                        st.markdown(f"""
<div style="background:#FFFFFF; border:1px solid #E5E7EB; border-radius:8px; padding:14px;">
    <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:6px;">
        <span style="font-weight:700; color:#1F4E79; font-size:0.9rem;">{item['category']}</span>
        <span style="background:{status_color}; color:white; padding:2px 10px; border-radius:12px; font-size:0.72rem; font-weight:700;">{item['status'].upper()}</span>
    </div>
    <div style="font-size:0.85rem; color:#111827; margin-bottom:6px;">{item['action']}</div>
    <div style="display:flex; gap:20px; font-size:0.78rem; color:#6B7280;">
        <span>Due: <strong>{item['due_date']}</strong></span>
        <span>Buyer: <strong>{item['buyer_owner']}</strong></span>
        <span>Supplier: <strong>{item['supplier_owner']}</strong></span>
    </div>
    {f'<div style="font-size:0.78rem; color:#6B7280; margin-top:4px;">Notes: {item["notes"]}</div>' if item["notes"] else ''}
</div>""", unsafe_allow_html=True)
                    with col_btn:
                        st.markdown("<br>", unsafe_allow_html=True)
                        if st.button("✏️ Edit", key=f"edit_btn_{i}", use_container_width=True):
                            st.session_state.editing_item = i
                            st.rerun()

            # Export action items
            action_csv = io.StringIO()
            writer = csv.DictWriter(action_csv, fieldnames=["category","action","status","due_date","buyer_owner","supplier_owner","notes"])
            writer.writeheader()
            writer.writerows(st.session_state.action_items)
            st.download_button(
                label="📥 Download Action Items (CSV)",
                data=action_csv.getvalue(),
                file_name=f"action_items_{supplier_name.replace(' ','_')}_{period.replace(' ','_')}.csv",
                mime="text/csv",
                use_container_width=True
            )
        else:
            st.info("No action items added yet. Use the form above to add items from your review meeting.")


if __name__ == "__main__":
    main()